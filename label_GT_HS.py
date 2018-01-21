import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook


def onclick(event):
    global ix, iy
    
    ix, iy = event.xdata, event.ydata
    # print 'x = %d, y = %d'%(
    #     ix, iy)

    global coords
    global curr_pos
    global subj
    global ang
    coords.append((subj, curr_pos+1, int(round(ix)), int(round(iy))))
    print coords

    if len(coords) == 5:
        fig.canvas.mpl_disconnect(cid)

    return coords
    

def key_event(e):
    global curr_pos
    cid = 0

    if e.key == "right":
        curr_pos = curr_pos + 1
    elif e.key == "left":
        curr_pos = curr_pos - 1
    else:
        return
    curr_pos = curr_pos % len(plots)

    ax.cla()
    ax.imshow(plots[curr_pos])
    cid = fig.canvas.mpl_connect('button_press_event',onclick)
    fig.canvas.draw()


cid = 0
curr_pos = 0


img_path = "/Volumes/2TB/IET_CV/HSdata/CASIA-B_054/frame"
wb_path = "/Volumes/2TB/IET_CV/excel/groundTruth/CASIA-B_054.xlsx"
wb = load_workbook(wb_path)

for roots, dirs, files in os.walk(img_path):
    # print roots
    folders = roots.split('/')
    # print folders

    if len(folders) == 8:
        all_frames = filter(lambda x: x.lower().endswith(('.png')),
            files)
        # print all_frames
        subj = folders[-1]
        plots = []
        coords = []
        for frame in all_frames:
            # print frame
            img_path = os.path.join(roots, frame)
            img = plt.imread(img_path)
            plots.append(img)
        fig = plt.gcf()
        ax = plt.gca()
        fig.canvas.mpl_connect('key_press_event', key_event)
        # print plots
        ax.imshow(plots[0])
        plt.show()
    
        for HS in coords:
            ws_name = subj
            if ws_name in wb.sheetnames:
                ws = wb[ws_name]
            else:
                ws = wb.create_sheet(0, ws_name)

            ws['A'+str(HS[1])] = HS[2]
            ws['B'+str(HS[1])] = HS[3]
        
        wb.save(wb_path)